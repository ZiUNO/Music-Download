# -*- coding: utf-8 -*-
"""
* @Author: ziuno
* @Software: PyCharm
* @Time: 2018/10/17 21:25
"""

import abc
import os
from random import randint
from time import sleep

import openpyxl
import requests
import xlrd


class Music(object):

    def __del__(self):
        self._data.clear()
        self._music_name = None
        self._singer_name = None

    def __init__(self):
        self._music_name = None
        self._data = []

    @property
    def music_list(self):
        music_list = []
        for row in self._data:
            music_list.append('%s-%s' % (row[0], row[1]))
        return list(set(music_list))

    @staticmethod
    def download(name_to_download, path=r'data\source.xlsx'):
        print('Begin to download %s' % name_to_download)
        wb = xlrd.open_workbook(path)
        ws = wb.sheet_by_name('music')
        name_to_download = Music.__handle_name(name_to_download)
        for row_index in range(1, ws.nrows):
            row = ws.row(row_index)
            name = '%s-%s' % (row[0].value, row[1].value)
            name = Music.__handle_name(name)
            url = row[2].value
            if name != name_to_download:
                continue
            if os.path.exists(r'Downloads\%s.m4a' % name):
                print('%s.m4a already exists.' % name)
                break
            print('Downloading %s...' % name, end='')
            while True:
                try:
                    music = requests.get(url, timeout=10)
                    print("successful")
                    break
                except:
                    print('failed, please wait a minute...')
                    sleep(randint(30, 50))
                    print('Downloading %s again...' % name, end='')
            with open(r'Downloads\%s.m4a' % name, 'wb') as file:
                file.write(music.content)
            break

    @abc.abstractmethod
    def search(self):
        pass

    def _write_search_log(self):
        print('Writing save source log...', end='')
        wb = openpyxl.load_workbook(r'data\source.xlsx')
        ws = wb['log']
        singer_name = self._singer_name
        if singer_name is None:
            singer_name = '#UNSET'
        ws.append([self._music_name, singer_name])
        wb.save(r'data\source.xlsx')
        wb.close()
        print('Done.')

    def _have_searched(self):
        wb = xlrd.open_workbook(r'data\source.xlsx')
        ws = wb.sheet_by_name('log')
        for row_index in range(1, ws.nrows):
            if self._music_name == ws.row(row_index)[0].value:
                if ws.row(row_index)[1].value == '#UNSET' or \
                        (self._singer_name is not None and ws.row(row_index)[1].value == self._singer_name):
                    print('%s is in the search log.' % self._music_name)
                    self._load_search_history()
                    return True
        return False

    def _load_search_history(self, path=r'data\source.xlsx'):
        if len(self._data) != 0:
            print('Have loaded search history.')
            return
        print('Loading search history from source.xlsx...', end='')
        wb = xlrd.open_workbook(path)
        ws = wb.sheet_by_name('music')
        for row_index in range(1, ws.nrows):
            if self._music_name in ws.row(row_index)[0].value:
                if self._singer_name is not None and (self._singer_name not in ws.row(row_index)[1].value):
                    continue
                name = ws.row(row_index)[0].value
                singer_name = ws.row(row_index)[1].value
                url = ws.row(row_index)[2].value
                self._data.append([name, singer_name, url])
        print('Done.')

    @property
    def singer_name(self):
        return self._singer_name

    @singer_name.setter
    def singer_name(self, name):
        if len(name) == 0:
            self._singer_name = None
        else:
            self._singer_name = name

    @property
    def music_name(self):
        return self._music_name

    @music_name.setter
    def music_name(self, name):
        self._music_name = name

    def save_source(self):
        if len(self._data) == 0:
            print('暂无数据或已保存资源')
        else:
            print('Saving source...')
            if self._have_searched():
                return
            wb = openpyxl.load_workbook(r'data\source.xlsx')
            ws = wb['music']
            for row_index in range(len(self._data)):
                ws.append(self._data[row_index])
            wb.save(r'data\source.xlsx')
            wb.close()
            print('Source save complete.')
            self._write_search_log()

    @staticmethod
    def clear_history():
        wb = openpyxl.Workbook()
        ws_music = wb.create_sheet('music')
        ws_log = wb.create_sheet('log')
        ws_music.cell(1, 1).value = '# COMMIT MUSIC_NAME SINGER_NAME SOURCE_LINK'

        ws_log.cell(1, 1).value = '# COMMIT SOURCE SEARCH LOG'
        if os.path.exists(r'data\source.xlsx'):
            try:
                os.remove(r'data\source.xlsx')
            except PermissionError:
                print('请先关闭已打开的source.xlsx文件或解除占用')
                os.system('pause')
                os.system('cls')
                return False
        wb.save(r'data\source.xlsx')
        wb.close()
        return True

    @staticmethod
    def __handle_name(name):
        name = list(name)
        illegal_characters = ['\\', '/', ':', '*', '?', '"', '<', '>', '|']
        for index in range(len(name)):
            if name[index] in illegal_characters:
                name[index] = '_'
        return ''.join(name)

    @staticmethod
    def download_all(path=r'data\source.xlsx'):
        print('Begin to download...')
        wb = xlrd.open_workbook(path)
        ws = wb.sheet_by_name('music')
        sleep_count = randint(5, 10)
        for row_index in range(1, ws.nrows):
            row = ws.row(row_index)
            name = '%s-%s' % (row[0].value, row[1].value)
            name = Music.__handle_name(name)
            url = row[2].value
            if os.path.exists(r'Downloads\%s.m4a' % name):
                print('%s.m4a already exists.' % name)
                continue
            print('Downloading %s...' % name, end='')
            while True:
                try:
                    music = requests.get(url, timeout=10)
                    print("successful")
                    break
                except BaseException:
                    print('failed, please wait a minute...')
                    sleep(randint(30, 50))
                    print('Downloading %s again...' % name, end='')
            with open(r'Downloads\%s.m4a' % name, 'wb') as file:
                file.write(music.content)
            sleep(randint(1, 3))
            sleep_count -= 1
            if sleep_count == 0:
                sleep_time = randint(10, 50)
                print('Sleep %d seconds...' % sleep_time, end='')
                sleep(sleep_time)
                print('Done.')
                print('Begin to download again...')
                sleep_count = randint(5, 10)
        print('Download complete')
